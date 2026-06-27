import tkinter as tk
from tkinter import filedialog, messagebox
from search_engine import search_song
from openpyxl import load_workbook
import csv
import threading


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Aikatsu Song DB Generator")
        self.root.geometry("650x450")

        self.file_path = ""

        # UI
        self.btn_select = tk.Button(root, text="Excelファイル選択", command=self.select_file)
        self.btn_select.pack(pady=10)

        self.label = tk.Label(root, text="未選択")
        self.label.pack()

        self.btn_run = tk.Button(root, text="実行", command=self.run_thread, bg="lightblue")
        self.btn_run.pack(pady=10)

        self.log = tk.Text(root, height=20)
        self.log.pack(fill=tk.BOTH, expand=True)

    # -------------------------
    # ファイル選択
    # -------------------------
    def select_file(self):
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")]
        )

        if self.file_path:
            self.label.config(text=self.file_path)
            self.log_insert("ファイル選択完了")

    # -------------------------
    # スレッド起動
    # -------------------------
    def run_thread(self):
        if not self.file_path:
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return

        thread = threading.Thread(target=self.run, daemon=True)
        thread.start()

    # -------------------------
    # メイン処理
    # -------------------------
    def run(self):
        self.log_insert("処理開始...")

        wb = load_workbook(self.file_path)
        ws = wb.active

        results = []

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        total = len(rows)

        for i, row in enumerate(rows, start=1):
            title = row[0]
            if not title:
                continue

            self.log_insert(f"[{i}/{total}] {title} → 処理中")

            try:
                song = search_song(title, f"AKT-{i:06d}")
                results.append(song.to_dict())
                self.log_insert(f"→ OK: {title}")
            except Exception as e:
                self.log_insert(f"→ ERROR: {title} ({e})")

        if results:
            output = self.file_path.replace(".xlsx", "_output.csv")

            with open(output, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)

            self.log_insert(f"完了: {output}")
            messagebox.showinfo("完了", "CSV生成が完了しました")

        else:
            self.log_insert("結果なし")

    # -------------------------
    # ログ表示
    # -------------------------
    def log_insert(self, text):
        self.log.insert(tk.END, text + "\n")
        self.log.see(tk.END)
        self.root.update_idletasks()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()