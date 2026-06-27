from openpyxl import load_workbook
from search_engine import search_song
import csv
import os

def run_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    songs = []

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        title = row[0]

        if not title:
            continue

        print(f"[{i}] {title}")

        song = search_song(title, f"AKT-{i:06d}")
        songs.append(song.to_dict())

    return songs


def save_csv(songs, output_path="output.csv"):
    keys = songs[0].keys()

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(songs)

    print(f"CSV保存完了: {output_path}")


if __name__ == "__main__":
    path = input("Excelファイルのパスを入力してください: ")
    songs = run_excel(path)
    save_csv(songs)